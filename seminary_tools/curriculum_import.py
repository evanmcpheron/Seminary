from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .common import sha256_file, slugify

EXPECTED_SHEETS = {
    "Wheaton Plan": ("Wheaton College", 0),
    "GCTS Plan": ("Gordon-Conwell Theological Seminary", 4),
    "Princeton Plan": ("Princeton Theological Seminary", 7),
}


def _project_year(sheet_name: str, source_year: str) -> int:
    if source_year.lower().startswith("pre-matriculation"):
        return 8
    digits = "".join(ch for ch in source_year if ch.isdigit())
    if not digits:
        raise ValueError(f"Cannot map project year from {source_year!r} in {sheet_name}")
    local_year = int(digits)
    offset = EXPECTED_SHEETS[sheet_name][1]
    return local_year + offset


def _term_id(sheet_name: str, term: str) -> str:
    normalized = term.strip().lower()
    if normalized == "before year 1":
        return "pre-matriculation"
    if sheet_name == "Wheaton Plan":
        mapping = {"fall": "term-01-fall", "spring": "term-02-spring", "summer": "term-03-summer"}
    elif sheet_name == "GCTS Plan":
        mapping = {"fall": "term-01-fall", "january": "term-02-january", "spring": "term-03-spring", "summer": "term-04-summer"}
    else:
        mapping = {"fall": "term-01-fall", "spring": "term-02-spring", "summer": "term-03-summer"}
    if normalized in mapping:
        return mapping[normalized]
    if "milestone" in normalized:
        return "milestones"
    return slugify(term)


def _record_type(sheet_name: str, code: str, credits: Any, status: str) -> str:
    status_l = status.lower()
    code_l = code.lower()
    if sheet_name != "Princeton Plan" and isinstance(credits, (int, float)):
        return "course"
    if "exam" in status_l or "examination" in code_l:
        return "examination"
    if "competency" in status_l or "language" in code_l or "preparation" in status_l:
        return "competency"
    if "seminar" in status_l or "doctoral seminar" in code_l or "core area" in code_l:
        return "doctoral-seminar-category"
    return "milestone"


def import_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    records: list[dict[str, Any]] = []

    for sheet_name, (institution, _) in EXPECTED_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Missing required worksheet: {sheet_name}")
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[4]]
        normalized_headers = [str(value).strip() if value is not None else "" for value in headers]
        expected_prefix = ["Year", "Term", "Course Code / Requirement", "Exact Current Catalog Title / Published Label", "Credits"]
        if normalized_headers[:5] != expected_prefix:
            raise ValueError(f"Unexpected headers in {sheet_name}: {normalized_headers[:5]!r}")

        for row_number in range(5, sheet.max_row + 1):
            values = [sheet.cell(row=row_number, column=column).value for column in range(1, 12)]
            if not any(value is not None for value in values):
                continue
            year, term, code, title, credits, requirement_status, description, rationale, prerequisites, note, source_url = values
            if not year or not term or not title:
                continue
            if str(year).startswith("Total") or str(title).startswith("Total"):
                continue

            record_type = _record_type(sheet_name, str(code or ""), credits, str(requirement_status or ""))
            project_year = _project_year(sheet_name, str(year))
            planned_term = _term_id(sheet_name, str(term))
            base_id = f"{project_year:02d}-{slugify(str(code or title))}-{slugify(str(title))}"
            record = {
                "record_id": base_id,
                "record_type": record_type,
                "institution": institution,
                "source_sheet": sheet_name,
                "source_row": row_number,
                "source_year": str(year),
                "source_term": str(term),
                "planned_year": project_year,
                "planned_term": planned_term,
                "course_code": str(code).strip() if code else None,
                "title": str(title).strip(),
                "credits": credits,
                "requirement_status": str(requirement_status or "").strip(),
                "catalog_description": str(description or "").strip(),
                "selection_rationale": str(rationale or "").strip(),
                "prerequisite_note": str(prerequisites or "").strip(),
                "offering_accuracy_note": str(note or "").strip(),
                "catalog_url": str(source_url or "").strip(),
            }
            records.append(record)

    numeric_credits = sum(float(r["credits"]) for r in records if isinstance(r.get("credits"), (int, float)))
    return {
        "schema_version": "1.0.0",
        "status": "imported-unreconciled",
        "source_spreadsheet": {
            "path": path.as_posix(),
            "sha256": sha256_file(path),
            "imported_at": datetime.now(timezone.utc).isoformat(),
            "catalog_basis": "Owner-approved 2026-27 curriculum spreadsheet",
        },
        "program": {
            "title": "Independent Theological Education Curriculum",
            "total_project_years": 12,
            "canonical_course_count": sum(1 for r in records if r["record_type"] == "course"),
            "imported_numeric_credits": numeric_credits,
        },
        "records": records,
    }
